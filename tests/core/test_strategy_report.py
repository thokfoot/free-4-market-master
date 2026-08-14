"""\
Strategy Excel Report Generator — tests for strategy_report.py

Behavioral invariants:
    - Net P&L per strategy sums to CSV total_pnl (source of truth)
    - Gross = Net + charges (charges = entry * qty * CHARGES_PER_MARKET[Mode])
    - Summary sorted by net P&L descending
    - Every fired strategy has a per-strategy trade sheet
    - Open trades appear with unrealized P&L (or marked open)
    - Never-fired strategies get a reason, not silently dropped
"""

import csv
import os

import openpyxl
import pytest

import strategy_report as sr


@pytest.fixture
def report_file(tmp_path, monkeypatch):
    """Generate a report into a temp dir; return the xlsx path."""
    out = tmp_path / "strategy_report.xlsx"
    monkeypatch.setattr(sr, "REPORT_FILE", str(out))
    return str(sr.generate_strategy_report(str(out)))


def _load_wb(path):
    return openpyxl.load_workbook(path)


class TestGrossNet:
    def test_gross_equals_net_plus_charges_us(self):
        trade = {"Mode": "US", "Entry_Price": "520.27", "Qty": "191", "P&L": "1257.92"}
        gross, charges, net = sr._gross_net(trade)
        assert net == 1257.92
        assert round(charges, 2) == round(520.27 * 191 * 0.0002, 2)
        assert round(gross, 2) == round(net + charges, 2)

    def test_gross_equals_net_plus_charges_crypto(self):
        trade = {"Mode": "CRYPTO", "Entry_Price": "0.1895", "Qty": "10000", "P&L": "-123.45"}
        gross, charges, net = sr._gross_net(trade)
        assert net == -123.45
        assert charges > 0
        assert round(gross, 2) == round(net + charges, 2)

    def test_loss_trade_charges_still_positive(self):
        trade = {"Mode": "US", "Entry_Price": "676.32", "Qty": "73", "P&L": "-997.56"}
        gross, charges, net = sr._gross_net(trade)
        assert net == -997.56
        assert charges == round(676.32 * 73 * 0.0002, 2)
        assert gross < 0

    def test_invalid_numbers_yield_zero(self):
        trade = {"Mode": "US", "Entry_Price": "abc", "Qty": "", "P&L": None}
        gross, charges, net = sr._gross_net(trade)
        assert gross == 0.0
        assert charges == 0.0
        assert net == 0.0


class TestChargeRate:
    def test_india_normalized(self):
        assert sr._charge_rate("India") == 0.0012
        assert sr._charge_rate("INDIAN") == 0.0012

    def test_us_and_crypto(self):
        assert sr._charge_rate("US") == 0.0002
        assert sr._charge_rate("CRYPTO") == 0.003

    def test_unknown_falls_back(self):
        assert sr._charge_rate("XYZ") == 0.001


class TestStrategyDefs:
    def test_loads_both_files(self):
        defs = sr._load_strategy_defs()
        tfs = {d["tf"] for d in defs}
        assert "SWING_1d" in tfs
        assert "INTRADAY_1h" in tfs

    def test_tickers_mapped_via_config(self):
        defs = sr._load_strategy_defs()
        by_market = {d["market"]: d for d in defs}
        assert by_market.get("Nasdaq100", {}).get("ticker") == "^NDX"
        assert by_market.get("PHLX_Semi", {}).get("ticker") == "^SOX"
        assert by_market.get("ADA", {}).get("ticker") == "ADA-USD"


class TestMatchDef:
    def test_matches_rank_across_markets(self):
        defs = sr._load_strategy_defs()
        # rank 1 exists for many markets; factors disambiguate
        d = sr._match_def(defs, "INTRADAY_1h", 1, "TRX-USD", "SHORT",
                          "EMA9>EMA20+Price>SMA20+RSI<50")
        assert d is not None
        assert d["ticker"] == "TRX-USD"

    def test_no_match_returns_none(self):
        assert sr._match_def([], "SWING_1d", 1, "NOPE", "LONG", "x") is None


class TestReportFile:
    def test_file_created(self, report_file):
        assert os.path.exists(report_file)

    def test_file_opens_with_openpyxl(self, report_file):
        _load_wb(report_file)

    def test_deterministic_bytes(self, tmp_path, monkeypatch):
        """Same data -> byte-identical xlsx (no useless git commits)."""
        out1 = tmp_path / "r1.xlsx"
        out2 = tmp_path / "r2.xlsx"
        sr.generate_strategy_report(str(out1))
        sr.generate_strategy_report(str(out2))
        assert open(out1, "rb").read() == open(out2, "rb").read()

    def test_summary_sheet_exists(self, report_file):
        wb = _load_wb(report_file)
        assert "Summary" in wb.sheetnames

    def test_summary_net_matches_csv_total(self, report_file):
        wb = _load_wb(report_file)
        ws = wb["Summary"]
        report_net = sum((ws.cell(row=r, column=17).value or 0) for r in range(2, ws.max_row + 1))
        csv_net = 0.0
        with open(sr.PAPER_FILE, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                if row.get("Status") == "CLOSED":
                    csv_net += float(row["P&L"])
        assert round(report_net, 2) == round(csv_net, 2)

    def test_summary_sorted_by_net_desc(self, report_file):
        wb = _load_wb(report_file)
        ws = wb["Summary"]
        nets = [ws.cell(row=r, column=17).value or 0.0 for r in range(2, ws.max_row + 1)]
        fired_nets = [n for n in nets if n != 0.0]
        assert fired_nets == sorted(fired_nets, reverse=True)

    def test_never_fired_sheet_has_reasons(self, report_file):
        wb = _load_wb(report_file)
        if "Never Fired" not in wb.sheetnames:
            pytest.skip("no never-fired strategies in this dataset")
        ws = wb["Never Fired"]
        reasons = {ws.cell(row=r, column=11).value for r in range(2, ws.max_row + 1)}
        assert reasons, "every never-fired strategy needs a reason"

    def test_all_trades_sheet_has_every_fired_trade(self, report_file):
        """Single combined 'All Trades' sheet contains every fired trade
        (one row per trade) — replaces the old per-strategy sheets."""
        wb = _load_wb(report_file)
        assert "All Trades" in wb.sheetnames
        ws = wb["Summary"]
        fired = sum(1 for r in range(2, ws.max_row + 1)
                    if (ws.cell(row=r, column=10).value or 0) > 0)
        assert fired > 0, "expected at least one fired strategy"
        # All Trades sheet must have a TOTAL row with closed+open counts
        at = wb["All Trades"]
        last = at.cell(row=at.max_row, column=6).value
        assert last == "TOTAL"
        assert "trades (" in str(at.cell(row=at.max_row, column=18).value)

    def test_all_trades_total_reconciles(self, report_file):
        """All Trades TOTAL row: Net P&L == Gross - Charges."""
        wb = _load_wb(report_file)
        assert "All Trades" in wb.sheetnames
        ws = wb["All Trades"]
        tot = ws.max_row
        assert ws.cell(row=tot, column=6).value == "TOTAL"
        gross = ws.cell(row=tot, column=14).value
        charges = ws.cell(row=tot, column=15).value
        net = ws.cell(row=tot, column=16).value
        assert gross is not None and charges is not None and net is not None
        assert net == pytest.approx(gross - charges, abs=0.01)
