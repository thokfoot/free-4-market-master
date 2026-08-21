"""
FREE 3-Market v5.10 — STRATEGY EXCEL REPORT GENERATOR
=====================================================
Builds logs/strategy_report.xlsx from live data:

  1. Summary  — every strategy (fired + never-fired), sorted by net P&L desc
  2. Per-strategy sheets — every trade with GROSS, CHARGES, NET P&L + subtotals
  3. Not Fired — strategies with 0 trades + why (from daily scan logs)

Auto-refresh: call generate_strategy_report() after each scan (bot.py main())
and after each live P&L check (live_pnl_updater.py main()).
"""

import glob
import hashlib
import json
import os
import subprocess
import zipfile as _zf

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

LOG_DIR = config.LOG_DIR if hasattr(config, "LOG_DIR") else os.path.join(os.path.dirname(__file__), "logs")
PAPER_FILE = os.path.join(LOG_DIR, "paper_trades.csv")
SWING_FILE = config.STRATEGY_FILE
INTRADAY_FILE = config.INTRADAY_STRATEGY_FILE
REPORT_FILE = os.path.join(LOG_DIR, "strategy_report.xlsx")
REPORT_SOURCE_FILE = os.path.join(LOG_DIR, "strategy_report_source.json")

# ── Excel styles ──
HDR_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
TOT_FILL = PatternFill(fill_type="solid", fgColor="FFE699")
GRN = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED = PatternFill(fill_type="solid", fgColor="FFC7CE")
YEL = PatternFill(fill_type="solid", fgColor="FFF2CC")
GRY = PatternFill(fill_type="solid", fgColor="EDEDED")
NONE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
THIN = Border()
CENTER = Alignment(horizontal="center")
LEFT = Alignment(horizontal="left")
RIGHT = Alignment(horizontal="right")


def _resolve_ticker(csv_market: str) -> str:
    """Map CSV market name to yfinance ticker.

    Exact TICKER_MAP lookup first, then fuzzy substring match - the SAME
    logic the scanners use (scanner.get_yf_ticker). This resolves aliases
    like 'XLK_Tech' -> XLK and 'XLF_Fin' -> XLF that strategy_report used
    to leave unresolved (causing 'Unmatched Trades' in the report).
    """
    # Region placeholders are NOT tickers - pass through unchanged.
    # (Without this guard, 'INDIAN' would fuzzy-match 'DIA' inside it.)
    if csv_market.upper() in ("INDIAN", "CRYPTO", "US", "INDIA"):
        return csv_market
    if csv_market in config.TICKER_MAP:
        return config.TICKER_MAP[csv_market]
    for key in config.TICKER_MAP:
        if key.lower() in csv_market.lower():
            return config.TICKER_MAP[key]
    return csv_market


def _resolve_tf(file_default: str, csv_tf) -> str:
    """Map the CSV TF column to an internal TimeFrame.

    1m -> GAP_DOWN_1m, 1h -> INTRADAY_1h, 1d* -> SWING_1d.
    Falls back to the file default when the column is missing/unparseable.
    """
    tf = str(csv_tf or "").strip().lower()
    if tf.startswith("1m"):
        return "GAP_DOWN_1m"
    if tf.startswith("1h"):
        return "INTRADAY_1h"
    if tf.startswith("1d"):
        return "SWING_1d"
    return file_default


def _load_strategy_defs():
    """Return list of strategy defs from both CSV files.

    Each def: {tf, market, ticker, region, rank, direction, factors,
               backtest_wr, backtest_pnl, trades, cost_rt}
    """
    import csv as _csv

    defs = []
    for path, file_default in ((SWING_FILE, "SWING_1d"), (INTRADAY_FILE, "INTRADAY_1h")):
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8-sig", newline="") as f:
            for row in _csv.DictReader(f):
                try:
                    rank = int(float(row["Final_Rank"]))
                except (KeyError, ValueError, TypeError):
                    continue
                market = str(row.get("Market", "")).strip()
                ticker = _resolve_ticker(market)
                tf = _resolve_tf(file_default, row.get("TF"))
                region = str(row.get("Region", "")).strip().upper()
                if region == "INDIA":
                    region = "INDIAN"
                defs.append({
                    "tf": tf,
                    "market": market,
                    "ticker": ticker,
                    "region": region,
                    "rank": rank,
                    "direction": str(row.get("Direction", "LONG")).strip().upper(),
                    "factors": str(row.get("Factors", "")).strip(),
                    "backtest_wr": _f(row.get("AvgWin%")),
                    "backtest_pnl": _f(row.get("Net_TotalPnL%_After_Charges") or row.get("Net_Total%_After_Charges")),
                    "trades": _f(row.get("Trades")),
                    "cost_rt": _f(row.get("Cost%_per_trade_RT") or row.get("Cost%_RT")),
                })
    # ── NSE FADE strategy family (ranks 992-996, v5.14) — hardcoded defs so
    # the live trades appear properly in the report even though the strategies
    # are driven by scanner_fade.py instead of a strategies CSV.
    variants = getattr(config, "FADE_VARIANTS", None)
    if variants:
        for v in variants:
            defs.append({
                "tf": "FADE_1h",
                "market": "NSE",
                "ticker": "NSE",
                "region": "INDIAN",
                "rank": v["rank"],
                "direction": "SHORT",
                "factors": v.get("factors", "Fade"),
                "backtest_wr": v.get("win_rate", 0),
                "backtest_pnl": 4.15,
                "trades": v.get("trades_count", 0),
                "cost_rt": 0.10,
            })
    us_variants = getattr(config, "US_FADE_VARIANTS", None)
    if us_variants:
        for v in us_variants:
            defs.append({
                "tf": "US_FADE_5m",
                "market": "US",
                "ticker": "US",
                "region": "US",
                "rank": v["rank"],
                "direction": "SHORT",
                "factors": v.get("factors", "US Fade"),
                "backtest_wr": v.get("win_rate", 0),
                "backtest_pnl": 8.5,
                "trades": v.get("trades_count", 0),
                "cost_rt": 0.10,
            })
    long_variants = getattr(config, "LONG_BOUNCE_VARIANTS", None)
    if long_variants:
        for v in long_variants:
            defs.append({
                "tf": "LONG_BOUNCE_5m",
                "market": "NSE",
                "ticker": "NSE",
                "region": "INDIAN",
                "rank": v["rank"],
                "direction": "LONG",
                "factors": v.get("factors", "Long Bounce"),
                "backtest_wr": v.get("win_rate", 0),
                "backtest_pnl": 56.0,
                "trades": v.get("trades_count", 0),
                "cost_rt": 0.10,
            })
    ipo_variants = getattr(config, "IPO_VARIANTS", None)
    if ipo_variants:
        for v in ipo_variants:
            defs.append({
                "tf": "IPO_1d",
                "market": "NSE",
                "ticker": "NSE",
                "region": "INDIAN",
                "rank": v["rank"],
                "direction": v["direction"],
                "factors": v.get("factors", "IPO edge"),
                "backtest_wr": v.get("win_rate", 0),
                "backtest_pnl": 5.0,
                "trades": v.get("trades_count", 0),
                "cost_rt": 0.10,
            })
    elif getattr(config, "FADE_RANK", None):
        defs.append({
            "tf": "FADE_1h",
            "market": "NSE",
            "ticker": "NSE",
            "region": "INDIAN",
            "rank": config.FADE_RANK,
            "direction": "SHORT",
            "factors": "Fade: 1h +3.5% vol2.2x RSI65 prev-high-break (2yr OOS +4.15%/mo)",
            "backtest_wr": 41.61,
            "backtest_pnl": 4.15,
            "trades": 3028,
            "cost_rt": 0.10,
        })
    return defs


def _f(v):
    """Parse a numeric cell safely -> float or 0.0."""
    try:
        if v is None:
            return 0.0
        s = str(v).strip().replace("%", "").replace(",", "")
        if not s or s in ("nan", "NaN", "-"):
            return 0.0
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _load_trades():
    """Load paper_trades.csv -> list of dicts."""
    import csv as _csv

    rows = []
    if not os.path.exists(PAPER_FILE):
        return rows
    with open(PAPER_FILE, encoding="utf-8-sig", newline="") as f:
        for r in _csv.DictReader(f):
            rows.append(r)
    return rows


def _charge_rate(mode: str) -> float:
    mode_norm = str(mode or "US").upper()
    if mode_norm == "INDIA":
        mode_norm = "INDIAN"
    return config.CHARGES_PER_MARKET.get(mode_norm, 0.001)


def _gross_net(trade):
    """Return (gross, charges, net) for a CLOSED trade.

    CSV P&L is NET (charges already deducted by paper_trader).
    Gross = Net + charges, where charges = entry*qty*rate (same as paper_trader).
    """
    entry = _f(trade.get("Entry_Price"))
    qty = _f(trade.get("Qty"))
    rate = _charge_rate(trade.get("Mode"))
    charges = round(entry * qty * rate, 2)
    net = _f(trade.get("P&L"))
    gross = round(net + charges, 2)
    return gross, charges, net


def _latest_close():
    """Latest known close per ticker from daily_scan logs."""
    closes = {}
    scans = sorted(glob.glob(os.path.join(LOG_DIR, "daily_scan_*.json")))
    for path in scans:
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, ValueError):
            continue
        for run in data.get("runs", []):
            mc = run.get("market_close") or {}
            for tk, px in mc.items():
                try:
                    closes[tk] = float(px)
                except (ValueError, TypeError):
                    pass
    return closes


def _scan_reasons():
    """Aggregate 'why' info from daily_scan logs.

    Returns:
        scanned_tickers: set of tickers ever scanned
        fired: set of (rank, ticker, direction) that fired in any scan
        skipped: list of {ticker, direction, rank, reason}
    """
    scanned = set()
    fired = set()
    skipped = []
    for path in sorted(glob.glob(os.path.join(LOG_DIR, "daily_scan_*.json"))):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, ValueError):
            continue
        for run in data.get("runs", []):
            scanned.update(run.get("tickers_scanned") or [])
            for fp in run.get("fired_patterns") or []:
                try:
                    r = int(float(fp.get("rank", 0)))
                except (ValueError, TypeError):
                    r = 0
                fired.add((r, fp.get("ticker"), str(fp.get("direction", "")).upper()))
            for se in run.get("skipped_entries") or []:
                try:
                    r = int(float(se.get("rank", 0)))
                except (ValueError, TypeError):
                    r = 0
                skipped.append({
                    "ticker": se.get("ticker"),
                    "direction": str(se.get("direction", "")).upper(),
                    "rank": r,
                    "reason": se.get("reason", "Skipped"),
                })
    return scanned, fired, skipped


def _strategy_key(defn):
    return (defn["tf"], defn["rank"], defn["ticker"], defn["direction"])


def _match_def(defs, tf, rank, ticker, direction, factors):
    """Match a trade to its strategy definition (rank repeats across markets).

    Prefer exact (tf, rank, ticker, direction, factors); fall back to rank+ticker.
    """
    exact = [d for d in defs if d["tf"] == tf and d["rank"] == rank
             and d["ticker"] == ticker and d["direction"] == direction
             and d["factors"] == factors]
    if len(exact) == 1:
        return exact[0]
    by_tk = [d for d in defs if d["tf"] == tf and d["rank"] == rank
             and d["ticker"] == ticker and d["direction"] == direction]
    if by_tk:
        return by_tk[0]
    # Gap-down strategies (997/998) have Market/Ticker = region placeholder
    # ('INDIAN') while trades carry real tickers (PFC.NS, ABFRL.NS...).
    # Match on (tf, rank, direction) so the 14 gap-down trades resolve.
    # Same for FADE family (ranks 992-996) — def ticker is 'NSE', trades
    # carry real tickers.
    _fade_ranks = {v["rank"] for v in getattr(config, "FADE_VARIANTS", [])}
    _usfade_ranks = {v["rank"] for v in getattr(config, "US_FADE_VARIANTS", [])}
    _long_ranks = {v["rank"] for v in getattr(config, "LONG_BOUNCE_VARIANTS", [])}
    if rank in (997, 998) or rank in _fade_ranks or rank in _usfade_ranks or rank in _long_ranks or \
            (getattr(config, "FADE_RANK", None) and rank == config.FADE_RANK):
        gd = [d for d in defs if d["tf"] == tf and d["rank"] == rank
              and d["direction"] == direction]
        if gd:
            return gd[0]
    return None


def _suffix(tf):
    """Strategy suffix used in labels AND sheet names (must stay in sync)."""
    if tf == "FADE_1h":
        return "FD"
    if tf == "LONG_BOUNCE_5m":
        return "LB"
    if tf == "US_FADE_5m":
        return "USFD"
    if tf == "INTRADAY_1h":
        return "ID"
    if tf == "GAP_DOWN_1m":
        return "GD"
    return "SW"


def _label(defn):
    suffix = _suffix(defn["tf"])
    return f"#{defn['rank']}{suffix} {defn['ticker']} {defn['direction']}"


def _write_sheet(wb, title, header, rows, widths, center_cols=(), num_cols=()):
    ws = wb.create_sheet(title)
    for c, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = THIN
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            if c in center_cols:
                cell.alignment = CENTER
            elif c in num_cols:
                cell.alignment = RIGHT
            if c in num_cols and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{max(1, len(rows) + 1)}"
    return ws


def generate_strategy_report(report_file=None):
    """Regenerate logs/strategy_report.xlsx. Returns the output path."""
    out = report_file or REPORT_FILE

    # Rebuild portfolio before reporting so the workbook and portfolio are
    # derived from the same authoritative paper_trades.csv snapshot.
    try:
        from paper_trader import rebuild_portfolio_from_csv
        rebuild_portfolio_from_csv()
    except Exception as exc:
        print(f"[StrategyReport] Portfolio rebuild warning: {exc}")

    defs = _load_strategy_defs()
    trades = _load_trades()
    closes = _latest_close()
    scanned, fired, skipped = _scan_reasons()

    # ── Group trades per strategy ──
    # strategy key = (tf, rank, ticker, direction) using the def if matched,
    # else derived from the trade row itself.
    groups = {}
    unmatched = 0
    for t in trades:
        tf = str(t.get("TimeFrame", "SWING_1d"))
        try:
            rank = int(float(t.get("Pattern_Rank", 0)))
        except (ValueError, TypeError):
            rank = 0
        ticker = str(t.get("Ticker", ""))
        direction = str(t.get("Direction", "LONG")).upper()
        factors = str(t.get("Pattern_Factors", "")).strip()
        defn = _match_def(defs, tf, rank, ticker, direction, factors)
        if defn is None:
            unmatched += 1
            defn = {
                "tf": tf, "rank": rank, "ticker": ticker, "region": str(t.get("Mode", "US")).upper(),
                "direction": direction, "factors": factors,
                "backtest_wr": _f(t.get("Expected_WinRate")), "backtest_pnl": 0.0,
                "trades": 0, "cost_rt": 0.0, "market": ticker,
            }
        key = (defn["tf"], defn["rank"], defn["ticker"], defn["direction"])
        groups.setdefault(key, {"defn": defn, "trades": []})["trades"].append(t)

    # ── Per-strategy aggregates ──
    summary_rows = []
    for key, grp in groups.items():
        d = grp["defn"]
        closed = [t for t in grp["trades"] if str(t.get("Status", "")).upper() == "CLOSED"]
        open_t = [t for t in grp["trades"] if str(t.get("Status", "")).upper() != "CLOSED"]
        gross_t = charges_t = net_t = 0.0
        wins = losses = 0
        for t in closed:
            g, c, n = _gross_net(t)
            gross_t += g
            charges_t += c
            net_t += n
            wins += 1 if n > 0 else 0
            losses += 1 if n < 0 else 0
        win_rate = round(wins / len(closed) * 100, 1) if closed else 0.0
        label = _label(d)
        summary_rows.append({
            "label": label, "defn": d,
            "total": len(grp["trades"]), "closed": len(closed), "open": len(open_t),
            "wins": wins, "losses": losses, "win_rate": win_rate,
            "gross": round(gross_t, 2), "charges": round(charges_t, 2),
            "net": round(net_t, 2),
            "avg_net": round(net_t / len(closed), 2) if closed else 0.0,
        })

    # ── Strategies never fired ──
    fired_keys = set()
    for d in defs:
        if (d["tf"], d["rank"], d["ticker"], d["direction"]) in groups:
            fired_keys.add((d["tf"], d["rank"], d["ticker"]))
    not_fired = []
    _fade_ranks_pre = {v["rank"] for v in getattr(config, "FADE_VARIANTS", [])}
    _fade_ranks_pre |= {v["rank"] for v in getattr(config, "US_FADE_VARIANTS", [])}
    _fade_ranks_pre |= {v["rank"] for v in getattr(config, "LONG_BOUNCE_VARIANTS", [])}
    if getattr(config, "FADE_RANK", None):
        _fade_ranks_pre.add(config.FADE_RANK)
    for d in defs:
        key = (d["tf"], d["rank"], d["ticker"], d["direction"])
        if key in groups:
            continue
        ticker = d["ticker"]
        rk = d["rank"]
        if rk in _fade_ranks_pre:
            # FADE family: def ticker is the 'NSE' placeholder; signals carry real
            # NSE tickers, so match on rank+direction only.
            if any(r == rk and str(dir_).upper() == str(d["direction"]).upper()
                   for (r, _, dir_) in fired):
                reason = "Fired in scan but never entered"
            else:
                reason = "Fade universe scanned - pattern never matched"
        elif ticker not in scanned:
            reason = "Ticker never scanned in any run"
        elif (rk, ticker, d["direction"]) in fired:
            sk = [s for s in skipped if s["ticker"] == ticker and s["rank"] == rk
                  and s["direction"] == d["direction"]]
            reason = sk[0]["reason"] if sk else "Fired in scan but never entered"
        else:
            reason = "Scanned but pattern never matched"
        not_fired.append({**d, "reason": reason})

    # ── Sort: fired by net desc, then never-fired ──
    fired_sum = [s for s in summary_rows if s["total"] > 0]
    fired_sum.sort(key=lambda s: (s["net"], s["total"]), reverse=True)
    _fade_ranks = {v["rank"] for v in getattr(config, "FADE_VARIANTS", [])}
    if getattr(config, "FADE_RANK", None):
        _fade_ranks.add(config.FADE_RANK)
    unfired = [d for d in not_fired if d["rank"] not in (997, 998)]
    unfired.sort(key=lambda d: d["backtest_wr"], reverse=True)

    # ══════════════ BUILD WORKBOOK ══════════════
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ──
    header = ["#", "Strategy", "TF", "Market/Ticker", "Region", "Direction", "Factors",
              "Backtest WR%", "Backtest NetPnL%", "Trades", "Open", "Wins", "Losses",
              "WinRate%", "Gross P&L", "Charges", "Net P&L", "Avg Net/Trade"]
    rows = []
    rank_idx = 1
    for s in fired_sum:
        d = s["defn"]
        rows.append([rank_idx, s["label"], d["tf"], d["ticker"], d.get("region", ""),
                     d["direction"], d["factors"], d["backtest_wr"], d["backtest_pnl"],
                     s["total"], s["open"], s["wins"], s["losses"], s["win_rate"],
                     s["gross"], s["charges"], s["net"], s["avg_net"]])
        rank_idx += 1
    for d in unfired:
        rows.append([rank_idx, f"#{d['rank']}{_suffix(d['tf'])} {d['ticker']} {d['direction']}",
                     d["tf"], d["ticker"], d.get("region", ""), d["direction"], d["factors"],
                     d["backtest_wr"], d["backtest_pnl"], 0, 0, 0, 0, 0.0, 0.0, 0.0, 0.0, 0.0])
        rank_idx += 1

    ws = _write_sheet(wb, "Summary", header, rows,
                      widths=[5, 26, 12, 13, 9, 9, 52, 10, 11, 8, 7, 7, 8, 9, 12, 11, 12, 12],
                      center_cols=(1, 3, 5, 6, 9, 10, 11, 12, 13, 14),
                      num_cols=(8, 15, 16, 17, 18))
    # Highlight fired profitable / losing
    for r in range(2, len(fired_sum) + 2):
        net_cell = ws.cell(row=r, column=17)
        net = net_cell.value or 0.0
        if net > 0:
            net_cell.fill = GRN
        elif net < 0:
            net_cell.fill = RED
    for r in range(len(fired_sum) + 2, len(rows) + 2):
        ws.cell(row=r, column=17).fill = GRY

    # ── Sheet 2: ALL TRADES (single full-data sheet, combined) ──
    # One row per trade with every column from paper_trades.csv PLUS
    # Gross/Charges/Net breakdown. This single sheet is what gets sent
    # to Telegram for independent verification by any AI / human.
    all_header = ["Date", "Time_IST", "Mode", "Ticker", "Direction", "TimeFrame",
                  "Entry", "Qty", "SL", "Target", "MaxHold", "Exit",
                  "Hold/Exit_Time", "Gross P&L", "Charges", "Net P&L", "P&L%",
                  "Status", "Pattern_Rank", "Expected_WinRate", "Pattern_Factors",
                  "Signal_Indicators", "Reason"]
    all_rows = []
    gr_tot = ch_tot = net_tot = 0.0
    n_open = 0
    for t in trades:
        g = c = n = 0.0
        pnl_lbl = 0.0
        if str(t.get("Status", "")).upper() == "CLOSED":
            g, c, n = _gross_net(t)
            gr_tot += g
            ch_tot += c
            net_tot += n
        else:
            n_open += 1
            entry = _f(t.get("Entry_Price"))
            qty = _f(t.get("Qty"))
            rate = _charge_rate(t.get("Mode"))
            c = round(entry * qty * rate, 2)
            cur = closes.get(t.get("Ticker"))
            if cur is not None:
                g = round((cur - entry) * qty, 2) if t.get("Direction") == "LONG" \
                    else round((entry - cur) * qty, 2)
                n = round(g - c, 2)
                pnl_lbl = f"~{n:.2f} (unrealized)"
        all_rows.append([t.get("Date"), t.get("Time_IST"), t.get("Mode"),
                         t.get("Ticker"), t.get("Direction"), t.get("TimeFrame"),
                         _f(t.get("Entry_Price")), _f(t.get("Qty")),
                         _f(t.get("SL")), _f(t.get("Target")), t.get("MaxHold"),
                         _f(t.get("Exit_Price")), t.get("Exit_Time"),
                         round(g, 2), round(c, 2), n if isinstance(n, str) else round(n, 2),
                         _f(t.get("P&L_%")), t.get("Status"),
                         t.get("Pattern_Rank"), t.get("Expected_WinRate"),
                         t.get("Pattern_Factors"), t.get("Signal_Indicators"),
                         t.get("Reason")])
    # sort: newest first
    def _sort_key(r):
        try:
            return str(r[0] or "") + " " + str(r[1] or "")
        except Exception:
            return ""
    all_rows.sort(key=_sort_key, reverse=True)
    all_rows.append(["", "", "", "", "", "TOTAL", "", "", "", "", "", "",
                     "", round(gr_tot, 2), round(ch_tot, 2), round(net_tot, 2),
                     "", f"{len(trades)} trades ({len(trades) - n_open} closed / {n_open} open)",
                     "", "", "", "", ""])
    all_ws = _write_sheet(wb, "All Trades", all_header, all_rows,
                          widths=[12, 12, 9, 12, 10, 13, 11, 8, 10, 10, 8, 11,
                                  18, 11, 10, 12, 9, 12, 9, 10, 45, 40, 50],
                          center_cols=(3, 4, 5, 6, 9, 10, 12, 18, 19),
                          num_cols=(7, 8, 9, 10, 14, 15, 16, 17))
    all_ws.freeze_panes = "A2"
    all_ws.auto_filter.ref = f"A1:{get_column_letter(len(all_header))}{len(all_rows)}"
    tot_row = len(all_rows) + 1
    for c in range(1, len(all_header) + 1):
        all_ws.cell(row=tot_row, column=c).fill = TOT_FILL
        all_ws.cell(row=tot_row, column=c).font = Font(bold=True)
    # color Net P&L column (16)
    for r in range(2, tot_row):
        cell = all_ws.cell(row=r, column=16)
        v = cell.value
        if isinstance(v, (int, float)):
            cell.fill = GRN if v > 0 else (RED if v < 0 else NONE_FILL)
    # color P&L% column (17)
    for r in range(2, tot_row):
        cell = all_ws.cell(row=r, column=17)
        v = cell.value
        if isinstance(v, (int, float)):
            cell.fill = GRN if v > 0 else (RED if v < 0 else NONE_FILL)
    if net_tot > 0:
        all_ws.cell(row=tot_row, column=16).fill = GRN
    elif net_tot < 0:
        all_ws.cell(row=tot_row, column=16).fill = RED

    # Deterministic metadata: pin created/modified to the latest trade date
    # so identical data yields byte-identical xlsx (no useless git commits).
    try:
        from datetime import datetime as _dt
        latest = "2026-01-01 00:00:00"
        for t in trades:
            dt = str(t.get("Date") or "").strip()
            if dt and dt > latest:
                latest = dt
        dt_obj = _dt.strptime(latest, "%Y-%m-%d")
        wb.properties.created = dt_obj
        wb.properties.modified = dt_obj
        _save_deterministic(wb, out, dt_obj)
    except (ValueError, TypeError):
        _save_deterministic(wb, out, None)
    _write_source_manifest(out, trades)
    print(f"[StrategyReport] Saved {out} ({len(fired_sum)} fired / {len(not_fired)} never-fired strategies)")
    return out


def _save_deterministic(wb, out, dt_obj):
    """Save workbook, then pin docProps/core.xml to a fixed timestamp.

    openpyxl stamps `modified` at save time, which makes byte output vary
    between runs. Rewrite core.xml with a deterministic timestamp instead.
    """
    import datetime as _dtm
    import re as _re

    tmp_out = out + ".tmp"
    wb.save(tmp_out)
    if dt_obj is None:
        os.replace(tmp_out, out)
        return
    stamp = dt_obj.strftime("%Y-%m-%dT%H:%M:%SZ")
    tmp = out + ".det"
    dt_tuple = (dt_obj.year, dt_obj.month, dt_obj.day, dt_obj.hour, dt_obj.minute, dt_obj.second)
    with _zf.ZipFile(tmp_out, "r") as zin, _zf.ZipFile(tmp, "w", _zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            item.date_time = dt_tuple
            if item.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                text = _re.sub(r"(<dcterms:created[^>]*>)[^<]*(</dcterms:created>)",
                               rf'\g<1>{stamp}\g<2>', text)
                text = _re.sub(r"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                               rf'\g<1>{stamp}\g<2>', text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp, out)
    os.remove(tmp_out)


def _write_source_manifest(report_file, trades):
    """Record the exact ledger snapshot used to build the workbook."""
    try:
        with open(PAPER_FILE, "rb") as f:
            ledger_hash = hashlib.sha256(f.read()).hexdigest()
        dates = [str(t.get("Date", "")).strip() for t in trades if t.get("Date")]
        try:
            commit = subprocess.check_output(
                ["git", "rev-parse", "HEAD"], cwd=os.path.dirname(__file__),
                text=True, stderr=subprocess.DEVNULL,
            ).strip()
        except Exception:
            commit = "unknown"
        manifest = {
            "report_file": os.path.abspath(report_file),
            "ledger_file": os.path.abspath(PAPER_FILE),
            "ledger_sha256": ledger_hash,
            "ledger_rows": len(trades),
            "ledger_latest_date": max(dates) if dates else "",
            "git_commit": commit,
        }
        manifest_file = os.path.splitext(report_file)[0] + "_source.json"
        tmp = manifest_file + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)
        os.replace(tmp, manifest_file)
    except Exception as exc:
        print(f"[StrategyReport] Source manifest warning: {exc}")


if __name__ == "__main__":
    generate_strategy_report()
